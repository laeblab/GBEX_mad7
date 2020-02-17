from django.apps import apps
from django.http import HttpResponse
from django.views.generic import TemplateView, View
from django.views.generic.base import TemplateResponseMixin, ContextMixin

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.writer.excel import save_virtual_workbook

from GBEX_app.helpers import model_to_list_list


class GBEXindex(TemplateView):
	template_name = 'GBEX_app/index.html'
	namespace = None

	def data_counts(self):
		all_numbers = {model.__name__: [model.objects.count(), 0] for model in apps.get_app_config('GBEX_app').get_models() if hasattr(model, "GBEX_Page")}

		divider = max([x[0] for x in all_numbers.values()])
		if divider != 0:
			all_numbers = {key: [value[0], value[0] / divider * 100] for key, value in all_numbers.items()}
		return all_numbers


class GBEXList(TemplateResponseMixin, ContextMixin, View):
	"""
	Convert an entire model into a string array
	"""
	template_name = "GBEX_app/list.html"
	model = None

	def get(self, request, *args, **kwargs):
		context = self.get_context_data(**kwargs)
		context['model_name'] = self.model.__name__
		context['model_order'] = self.model.order
		context['data'] = model_to_list_list(self.model.objects.all())

		return self.render_to_response(context)


class ExcelExportView(View):
	model = None

	def get(self, request, *args, **kwargs):
		if kwargs['rids']:
			# find alle model instances med et id i rids OG som er i namespace
			ids = kwargs['rids'].split(",")
			objects = self.model.objects.filter(id__in=ids)
		else:
			# return all model instances from namespace
			objects = self.model.objects.filter()

		# turn objects into a workbook
		data = model_to_list_list(objects)
		wb = Workbook()
		ws = wb.active
		ws.title = self.model.__name__

		ws.append(self.model.order)
		for row in data:
			ws.append(row)

		tab = Table(displayName="Table1", ref=f"A1:{get_column_letter(len(self.model.order))}{len(data)}")
		style = TableStyleInfo(
			name="TableStyleMedium9", showFirstColumn=True, showLastColumn=False,
			showRowStripes=True, showColumnStripes=False
		)
		tab.tableStyleInfo = style
		ws.add_table(tab)

		response = HttpResponse(
			save_virtual_workbook(wb),
			content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
		response['Content-Disposition'] = 'attachment; filename="export.xlsx"'
		return response
